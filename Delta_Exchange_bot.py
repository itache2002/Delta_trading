import requests
import pandas as pd
import websocket
import datetime
import json
from talib import abstract
import hashlib
import hmac
import numpy as np
from openpyxl import Workbook, load_workbook


class Delat():
    def __init__(self,api_key, api_secret,lev):
        self.api_key = api_key
        self.api_secret = api_secret
        self.leverage = lev
        self.EMA = 5
        self.BBL = 15
        self.BBSD = 3
        self.headers = {'Accept': 'application/json'}
        self.historydf = None
        self.update_data_df= None
        self.entry = 0
        self.exit = 0
        self.stoploss= 0
        self.takeprofit = 0
        self.crossover = False


    def History(self):
        current_timestamp = int(datetime.datetime.now().timestamp())
        last_100_candel = current_timestamp - 90000

        params = {
            'resolution': '15m',
            'symbol': 'BTCUSDT',
            'start':last_100_candel, 
            'end': current_timestamp   
        }

        r = requests.get('https://api.delta.exchange/v2/history/candles', params=params, headers=self.headers)
        if r.status_code == 200:
            json_data = r.json()
            df = pd.DataFrame(json_data['result'])
            df['time'] = pd.to_datetime(df['time'], unit='s')  
            df['time'] = df['time'].dt.tz_localize('UTC').dt.tz_convert('Asia/Kolkata').dt.strftime('%Y-%m-%d %H:%M:%S')
            self.historydf = df
            self.historydf = self.historydf.iloc[::-1].reset_index(drop=True)
            self.historydf = self.historydf.drop(self.historydf.index[-1])
            self.historydf['upper'], self.historydf['middleband'], self.historydf['lower'] = np.round(abstract.BBANDS(self.historydf['close'], timeperiod=15, nbdevup=3.0, nbdevdn=3.0, matype=0),1)
            self.historydf['EMA'] =np.round(abstract.EMA(self.historydf['close'], 5),1)
        else:
            print(f"Failed to fetch data. Status code: {r.status_code}")


    def generate_signature(self,secret, message):
        message = bytes(message, 'utf-8')
        secret = bytes(secret, 'utf-8')
        hash = hmac.new(secret, message, hashlib.sha256)
        return hash.hexdigest()

    def get_time_stamp(self):
        d = datetime.datetime.utcnow()
        epoch = datetime.datetime(1970, 1, 1)
        return str(int((d - epoch).total_seconds()))
    
    def Place_generate_signature(self,method, endpoint, payload):
        timestamp = self.get_time_stamp()
        signature_data = method + timestamp + endpoint + payload
        message = bytes(signature_data, 'utf-8')
        secret = bytes(self.api_secret, 'utf-8')
        hash = hmac.new(secret, message, hashlib.sha256)
        return hash.hexdigest(), timestamp

    # To get the product_id of the any symbol
    def get_product_id(self,symbol):
        
        url = f"https://api.delta.exchange/v2/products/{symbol}"
        response = requests.get(url, headers= self.headers)

        if response.status_code == 200:
            data = response.json()
            if data["success"]:
                print(data["result"]["id"])
                return data["result"]["id"]
            
            else:
                return "Could not retrieve data. Please check the product symbol."
        else:
            return f"Failed to retrieve data. HTTP Status code: {response.status_code}"
        
    def place_order(self,side, qty, product_id, order_type="market_order", price=None):
        method = 'POST'
        url = "https://api.delta.exchange/v2/orders"
        endpoint = "/v2/orders"
        query_string = ''

        params = {
            "order_type": order_type,
            "size": qty,
            "side": side,
            "product_id": product_id,
        }
        if price:
            params["limit_price"] = str(price)

        payload = json.dumps(params).replace('','')
        signature, timestamp = self.Place_generate_signature(method, endpoint, payload)

        headers = {
            'api-key': self.api_key,
            'timestamp': timestamp,
            'signature': signature,
            'User-Agent': 'rest-client',
            'Content-Type': 'application/json'
        }

        response = requests.post(url, data=payload, headers=headers)
        return response.json()
    


    def add_to_excel(self , timestamp ,entry_price, exit_price, stop_loss, take_profit):
        try:
              wb = load_workbook('BUY.xlsx')
              sheet = wb.active
        except FileNotFoundError:
              wb = Workbook()
              sheet = wb.active
              sheet.append(['Entry Time', 'Entry Price', 'Exit Price','stop_loss','take_profit'])

        for row in sheet.iter_rows(min_row=2, max_col=7):
            existing_entry = [cell.value for cell in row]
            new_entry = [entry_price, exit_price, stop_loss, take_profit]

            if existing_entry[1:] == new_entry:
                print("Duplicate entry found. Exiting without adding.")
                return

        sheet.append([timestamp, entry_price, exit_price,stop_loss,take_profit])
        wb.save('BUY.xlsx')
        print("Entry added successfully.")
            
    def Strategy(self):
        tail_2 = self.historydf.tail(2)
        tail_1 = self.historydf.tail(1)
        current_price = int(tail_1['close'].iloc[0])
        time_stamp = tail_1['time'].iloc[0]
        

 
        if (self.historydf['EMA'][len(self.historydf.index)-3] < self.historydf['middleband'][len(self.historydf.index)-3]) and (self.historydf['EMA'][len(self.historydf.index)-2] > self.historydf['middleband'][len(self.historydf.index)-2]):
            self.crossover = True

        # crossup Strategy
        if self.crossover and self.historydf['EMA'][len(self.historydf)-2] > self.historydf['middleband'][len(self.historydf.index)-2]:
            print("##################")
            print(" BUY Signal ")
            print("##################")
            self.entry = float(tail_1['close'].iloc[0])
            self.stoploss= int(self.entry  - 100)
            self.takeprofit = int(self.entry + 200)
            self.exit = self.stoploss
            print(f"The current timestamp{time_stamp}")
            print(f"The entry price is {self.entry}")
            print(f"The  price is exit  {self.exit }")
            print(f"The stoploss price is {self.stoploss}")
            print(f"The EMA :{tail_1['EMA'].iloc[0]}")
            print(f"The Previous EMA :{tail_2['EMA'].iloc[0]}")
            print(f"The middelband :{tail_1['middleband'].iloc[0]}")
            print(f"The Previous middelband :{tail_2['middleband'].iloc[0]}")

         # crossdown Strategy    
        if self.crossover and self.historydf['EMA'][len(self.historydf)-2] < self.historydf['middleband'][len(self.historydf.index)-2]:
            print("##################")
            print(" SELL Signal ")
            print("##################")
            self.entry = float(tail_1['close'].iloc[0])
            self.stoploss= int(self.entry  + 100)
            self.takeprofit = int(self.entry - 200)
            self.exit = self.stoploss
            print(f"The current timestamp{time_stamp}")
            print(f"The entry price is {self.entry}")
            print(f"The  price is exit  {self.exit }")
            print(f"The stoploss price is {self.stoploss}")
            print(f"The EMA :{tail_1['EMA'].iloc[0]}")
            print(f"The Previous EMA :{tail_2['EMA'].iloc[0]}")
            print(f"The middelband :{tail_1['middleband'].iloc[0]}")
            print(f"The Previous middelband :{tail_2['middleband'].iloc[0]}")

        


        if current_price == self.stoploss:
            print("The trade ended  with a loss")
            # Setting the exit to stoploss
            self.exit = self.stoploss
            #add the data to excel 
            self.add_to_excel(time_stamp,self.entry,self.exit,self.stoploss,self.takeprofit)

        if current_price == self.takeprofit:
            print("The trade ended  with a profit")
            # Setting the exit to the take profit
            self.exit = self.takeprofit
            #add the data to excel 
            self.add_to_excel(time_stamp,self.entry,self.exit,self.stoploss,self.takeprofit)
        


    def Update_data(self,data):
        if data['close'] is None or data['candle_start_time'] is None:
            print("Warning: close or candle_start_time value is None. Skipping Update_data.")
            return
        start_time = data['candle_start_time']
        open_data = float(data['open'])
        close_data = float(data['close'])
        high_data = float(data['high'])
        low_data = float(data['low'])
        volume_data = float(data['volume'])

        if np.isnan(close_data):  # Check if close_data is NaN
            print("Warning: close_data is NaN. Skipping Update_data.")
            return

        new_df = pd.DataFrame({
            'time': [pd.to_datetime(start_time / 1000000, unit='s')],
            'open': [open_data],
            'close': [close_data],
            'high': [high_data],
            'low': [low_data],
            'volume': [volume_data]
        })

        self.historydf = pd.concat([self.historydf, new_df], ignore_index=True)
        self.historydf['EMA'] = np.round(abstract.EMA(self.historydf['close'],5),1)
        self.historydf['upper'], self.historydf['middleband'], self.historydf['lower'] = np.round(abstract.BBANDS(self.historydf['close'], timeperiod=15, nbdevup=3.0, nbdevdn=0.0, matype=0),1)

        # print(self.historydf)
        self.Strategy()


    def on_message(self,ws, message):
        data = json.loads(message)
        # print(data)
        self.Update_data(data)

    def on_close(self,ws, close_status_code, close_msg):
        print("Websocket closed")

    def on_error(self,ws, error):
        print(f"Error: {error}")

    def on_open(self,ws):
        print("Websocket opened")
        print("Waiting for the Signal....")
    # Send authentication message
        timestamp = self.get_time_stamp()
        signature_data = f"GET{timestamp}/live"
        signature = self.generate_signature( self.api_secret, signature_data)

        auth_msg = {
            "type": "auth",
            "payload": {
                "api-key": self.api_key,
                "signature": signature,
                "timestamp": timestamp
            }
        }

        ws.send(json.dumps(auth_msg))  
        # Subscribe to channels
        ''' 
        if u want to chang the candle_stick time u can  change candlestick_1m for example:
        you want to 15min time framse  u chang it to candlestick_15m

        '''
        subscribe_msg = {
            "type": "subscribe",
            "payload": {
                "channels": [
                    {"name": "candlestick_15m", "symbols": ['BTCUSDT']},
                ]
            }
        }
        ws.send(json.dumps(subscribe_msg))


    def live_data(self):
      ws = websocket.WebSocketApp('wss://socket.delta.exchange', on_open=self.on_open, on_message=self.on_message, on_error=self.on_error, on_close=self.on_close)
      ws.run_forever()

api_key = 'VraFQTfHWk1w7Id1uC1pJvyKQ5AWy5'
api_secret = 'XJofp39iR4p7092qkzJdH9VPmoZmHtWv2x0huuJWscPerHtGqcf2Uo996JMj'


delta = Delat(api_key,api_secret,25)

delta.History()
delta.live_data()
# delta.get_product_id('BTCUSDT')
# delta.has_15_minutes_passed(1708197060000000)