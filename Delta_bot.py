import requests
import pandas as pd
import websocket
import datetime
import json
from talib import abstract
import hashlib
import hmac
import numpy as np
from openpyxl import Workbook, load_workbook


class Delat():
    def __init__(self,api_key, api_secret,lev):
        self.api_key = api_key
        self.api_secret = api_secret
        self.leverage = lev
        self.EMA = 5
        self.BBL = 15
        self.BBSD = 3.0
        self.headers = {'Accept': 'application/json'}
        self.historydf = None
        self.update_data_df= None
        self.entry = 0
        self.exit = 0
        self.stoploss= 0
        self.takeprofit = 0
        self.crossover = False
        self.Previous= None



    def History(self):
        current_timestamp = int(datetime.datetime.now().timestamp())
        last_100_candel = current_timestamp - 90000

        params = {
            'resolution': '1m',
            'symbol': 'BTCUSDT',
            'start':last_100_candel, 
            'end': current_timestamp   
        }

        r = requests.get('https://api.delta.exchange/v2/history/candles', params=params, headers=self.headers)
        if r.status_code == 200:
            json_data = r.json()
            df = pd.DataFrame(json_data['result'])
            df['time'] = pd.to_datetime(df['time'], unit='s')  
            # df['time'] = df['time'].dt.tz_localize('UTC').dt.tz_convert('Asia/Kolkata').dt.strftime('%Y-%m-%d %H:%M:%S')
            self.historydf = df
            self.historydf = self.historydf.iloc[::-1].reset_index(drop=True)
            self.historydf = self.historydf.drop(self.historydf.index[-1])
            self.historydf['upper'], self.historydf['middleband'], self.historydf['lower'] = np.round(abstract.BBANDS(self.historydf['close'], timeperiod=15, nbdevup=3.0, nbdevdn=3.0, matype=0),1)
            self.historydf['EMA'] =np.round(abstract.EMA(self.historydf['close'], 5),1)
            print(self.historydf)

        else:
            print(f"Failed to fetch data. Status code: {r.status_code}")


    def generate_signature(self,secret, message):
        message = bytes(message, 'utf-8')
        secret = bytes(secret, 'utf-8')
        hash = hmac.new(secret, message, hashlib.sha256)
        return hash.hexdigest()

    def get_time_stamp(self):
        d = datetime.datetime.utcnow()
        epoch = datetime.datetime(1970, 1, 1)
        return str(int((d - epoch).total_seconds()))
    
    def Place_generate_signature(self,method, endpoint, payload):
        timestamp = self.get_time_stamp()
        signature_data = method + timestamp + endpoint + payload
        message = bytes(signature_data, 'utf-8')
        secret = bytes(self.api_secret, 'utf-8')
        hash = hmac.new(secret, message, hashlib.sha256)
        return hash.hexdigest(), timestamp

    # To get the product_id of the any symbol
    def get_product_id(self,symbol):
        
        url = f"https://api.delta.exchange/v2/products/{symbol}"
        response = requests.get(url, headers= self.headers)

        if response.status_code == 200:
            data = response.json()
            if data["success"]:
                print(data["result"]["id"])
                return data["result"]["id"]
            
            else:
                return "Could not retrieve data. Please check the product symbol."
        else:
            return f"Failed to retrieve data. HTTP Status code: {response.status_code}"
        
    def place_order(self,side, qty, product_id, order_type="market_order", price=None):
        method = 'POST'
        url = "https://api.delta.exchange/v2/orders"
        endpoint = "/v2/orders"
        query_string = ''

        params = {
            "order_type": order_type,
            "size": qty,
            "side": side,
            "product_id": product_id,
        }
        if price:
            params["limit_price"] = str(price)

        payload = json.dumps(params).replace('','')
        signature, timestamp = self.Place_generate_signature(method, endpoint, payload)

        headers = {
            'api-key': self.api_key,
            'timestamp': timestamp,
            'signature': signature,
            'User-Agent': 'rest-client',
            'Content-Type': 'application/json'
        }

        response = requests.post(url, data=payload, headers=headers)
        return response.json()
    


    def add_to_excel(self , timestamp ,entry_price, exit_price, stop_loss, take_profit):
        try:
              wb = load_workbook('BUY.xlsx')
              sheet = wb.active
        except FileNotFoundError:
              wb = Workbook()
              sheet = wb.active
              sheet.append(['Entry Time', 'Entry Price', 'Exit Price','stop_loss','take_profit'])

        for row in sheet.iter_rows(min_row=2, max_col=7):
            existing_entry = [cell.value for cell in row]
            new_entry = [entry_price, exit_price, stop_loss, take_profit]

            if existing_entry[1:] == new_entry:
                print("Duplicate entry found. Exiting without adding.")
                return

        sheet.append([timestamp, entry_price, exit_price,stop_loss,take_profit])
        wb.save('BUY.xlsx')
        print("Entry added successfully.")
        

            
    def Strategy(self,start_time,open_data,close_data,high_data,low_data):
        current_price = int(close_data)
        print(f"The current EMA:{self.historydf['EMA'].iloc[-1]}")
        print(f"The current price : {current_price}")
        # crossup Strategy
        if (self.historydf['EMA'].iloc[-2] < self.historydf['middleband'].iloc[-2]) and (abs(int(self.historydf['middleband'].iloc[-1]) - int(self.historydf['EMA'].iloc[-1])) <= 1) :
            print("##################")
            print(" BUY Signal ")
            print("##################")
            self.entry = float(tail_1['close'].iloc[0])
            self.stoploss= int(self.entry  - 100)
            self.takeprofit = int(self.entry + 200)
            self.exit = self.stoploss
            print(f"The current timestamp{time_stamp}")
            print(f"The entry price is {self.entry}")
            print(f"The  price is exit  {self.exit }")
            print(f"The stoploss price is {self.stoploss}")


         # crossdown Strategy    
        if (self.historydf['EMA'].iloc[-2] > self.historydf['middleband'].iloc[-2]) and (abs(int(self.historydf['middleband'].iloc[-1]) - int(self.historydf['EMA'].iloc[-1])) <= 1) :
            print("##################")
            print(" SELL Signal ")
            print("##################")
            self.entry = float(tail_1['close'].iloc[0])
            self.stoploss= int(self.entry  + 100)
            self.takeprofit = int(self.entry - 200)
            self.exit = self.stoploss
            print(f"The current timestamp{time_stamp}")
            print(f"The entry price is {self.entry}")
            print(f"The  price is exit  {self.exit }")
            print(f"The stoploss price is {self.stoploss}")
        
        if current_price == self.stoploss:
            print("The trade ended  with a loss")
            # Setting the exit to stoploss
            self.exit = self.stoploss
            #add the data to excel 
            self.add_to_excel(start_time,self.entry,self.exit,self.stoploss,self.takeprofit)

        if current_price == self.takeprofit:
            print("The trade ended  with a profit")
            # Setting the exit to the take profit
            self.exit = self.takeprofit
            #add the data to excel 
            self.add_to_excel(start_time,self.entry,self.exit,self.stoploss,self.takeprofit)
        
    

    def calculate_candle_close_time(self, open_time):
            """
            Calculate the closing time of a 15-minute candle given its opening time.

            :param open_time: The opening time of the candle in UNIX timestamp format (milliseconds).
            :return: The closing time as a formatted string in the 'Asia/Kolkata' timezone.
            """
            # Convert the UNIX timestamp from microseconds to seconds
            open_time_seconds = open_time / 1e6

            open_time_dt = datetime.datetime.utcfromtimestamp(open_time_seconds)
            
            # Add 15 minutes to the opening time
            close_time_dt = open_time_dt + datetime.timedelta(minutes=1)


            return close_time_dt

    def is_candle_closed(self, open_time):
        """
        Check if the current time is past the closing time of a 15-minute candle.

        :param open_time: The opening time of the candle in UNIX timestamp format (milliseconds).
        :return: Boolean indicating whether the candle is closed (True) or not (False).
        """
        # Calculate the candle's closing time
        candle_close_time = self.calculate_candle_close_time(open_time)
        
        # Get the current time in UTC
        current_time_utc = datetime.datetime.utcnow()
       # Check if the current time is past the candle's closing time
        close = current_time_utc >= candle_close_time

        return current_time_utc >= candle_close_time
    
    def calculate_ema(self, prices, period):
        multiplier = 2 / (period + 1)
        return prices.ewm(alpha=multiplier, adjust=False).mean()



    def Update_data(self,data):
        if data['close'] is None or data['candle_start_time'] is None:
            print("Warning: close or candle_start_time value is None. Skipping Update_data.")
            return
        start_time = data['candle_start_time']
        open_data = float(data['open'])
        close_data = float(data['close'])
        high_data = float(data['high'])
        low_data = float(data['low'])
        volume_data = float(data['volume'])

        if np.isnan(close_data):  # Check if close_data is NaN
            print("Warning: close_data is NaN. Skipping Update_data.")
            return
        
        try:
            new_df = pd.DataFrame({
                'time': [pd.to_datetime(start_time / 1000000, unit='s')],
                'open': [open_data],
                'close': [close_data],
                'high': [high_data],
                'low': [low_data],
                'volume': [volume_data]
            })
            # Use pd.concat to add the new row to the existing DataFrame
            self.historydf = pd.concat([self.historydf, new_df], ignore_index=True)

            # if self.historydf['close'].iloc[-2] != close_data:
            #     print("True")
            # self.historydf['EMA'] = np.round(abstract.EMA(self.historydf['close'],5),1)
            # self.historydf['EMA'] = round(self.historydf['close'].ewm(span=5, adjust=False, min_periods=5).mean(),1)
            self.historydf['EMA'] = np.round(abstract.EMA(self.historydf['close'],5),1)
            upper, middle, lower = abstract.BBANDS(self.historydf['close'], timeperiod=15, nbdevup=3.0, nbdevdn=3.0, matype=0)
            # Assign each individually, rounding as necessary
            self.historydf['upper'] = np.round(upper, 1)
            self.historydf['middleband'] = np.round(middle, 1)
            self.historydf['lower'] = np.round(lower, 1)
            # self.historydf[['upper', 'middleband', 'lower']] = np.round(abstract.BBANDS(self.historydf['close'], timeperiod=15, nbdevup=3.0, nbdevdn=3.0, matype=0), 1)
            self.historydf.to_excel('live-data.xlsx')
            print(self.historydf)
        except ValueError as e:
            print(f"Error converting data types: {e}")
            return

        # print(self.historydf)
        self.Strategy(start_time,open_data,close_data,high_data,low_data)


    def on_message(self,ws, message):
        data = json.loads(message)
        # print(data)
        self.Update_data(data)

    def on_close(self,ws, close_status_code, close_msg):
        print("Websocket closed")

    def on_error(self,ws, error):
        print(f"Error: {error}")

    def on_open(self,ws):
        print("Websocket opened")
        print("Waiting for the Signal....")
    # Send authentication message
        timestamp = self.get_time_stamp()
        signature_data = f"GET{timestamp}/live"
        signature = self.generate_signature( self.api_secret, signature_data)

        auth_msg = {
            "type": "auth",
            "payload": {
                "api-key": self.api_key,
                "signature": signature,
                "timestamp": timestamp
            }
        }

        ws.send(json.dumps(auth_msg))  
        # Subscribe to channels
        ''' 
        if u want to chang the candle_stick time u can  change candlestick_1m for example:
        you want to 15min time framse  u chang it to candlestick_15m

        '''
        subscribe_msg = {
            "type": "subscribe",
            "payload": {
                "channels": [
                    {"name": "candlestick_1m", "symbols": ['BTCUSDT']},
                ]
            }
        }
        ws.send(json.dumps(subscribe_msg))


    def live_data(self):
      ws = websocket.WebSocketApp('wss://socket.delta.exchange', on_open=self.on_open, on_message=self.on_message, on_error=self.on_error, on_close=self.on_close)
      ws.run_forever()

api_key = 'VraFQTfHWk1w7Id1uC1pJvyKQ5AWy5'
api_secret = 'XJofp39iR4p7092qkzJdH9VPmoZmHtWv2x0huuJWscPerHtGqcf2Uo996JMj'


delta = Delat(api_key,api_secret,25)

delta.History()
delta.live_data()