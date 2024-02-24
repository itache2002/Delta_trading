import requests
import pandas as pd
import websocket
import datetime
import json
from talib import abstract
import hashlib
import hmac
import numpy as np
from openpyxl import Workbook, load_workbook


class Delat():
    def __init__(self,api_key, api_secret,lev):
        self.api_key = api_key
        self.api_secret = api_secret
        self.leverage = lev
        self.EMA = 5
        self.BBL = 15
        self.BBSD = 3.0
        self.headers = {'Accept': 'application/json'}
        self.historydf = None
        self.update_data_df= None
        self.entry = 0
        self.exit = 0
        self.stoploss= 0
        self.takeprofit = 0
        self.crossover = False
        self.finaldf=None
        self.livedf= None



    def History(self):
        current_timestamp = int(datetime.datetime.now().timestamp())
        last_100_candel = current_timestamp - 90000

        params = {
            'resolution': '1m',
            'symbol': 'BTCUSDT',
            'start':last_100_candel, 
            'end': current_timestamp   
        }

        r = requests.get('https://api.delta.exchange/v2/history/candles', params=params, headers=self.headers)
        if r.status_code == 200:
            json_data = r.json()
            df = pd.DataFrame(json_data['result'])
            df['time'] = pd.to_datetime(df['time'], unit='s')  
            # df['time'] = df['time'].dt.tz_localize('UTC').dt.tz_convert('Asia/Kolkata').dt.strftime('%Y-%m-%d %H:%M:%S')
            self.historydf = df
            self.historydf = self.historydf.iloc[::-1].reset_index(drop=True)
            # self.historydf = self.historydf.drop(self.historydf.index[-1])
            self.historydf['upper'], self.historydf['middleband'], self.historydf['lower'] = np.round(abstract.BBANDS(self.historydf['close'], timeperiod=15, nbdevup=3.0, nbdevdn=3.0, matype=0),1)
            self.historydf['EMA'] =np.round(abstract.EMA(self.historydf['close'], 5),1)
            self.finaldf = self.historydf
            self.livedf = self.historydf
            # print(self.finaldf)


        else:
            print(f"Failed to fetch data. Status code: {r.status_code}")


    def generate_signature(self,secret, message):
        message = bytes(message, 'utf-8')
        secret = bytes(secret, 'utf-8')
        hash = hmac.new(secret, message, hashlib.sha256)
        return hash.hexdigest()

    def get_time_stamp(self):
        d = datetime.datetime.utcnow()
        epoch = datetime.datetime(1970, 1, 1)
        return str(int((d - epoch).total_seconds()))
    
    def Place_generate_signature(self,method, endpoint, payload):
        timestamp = self.get_time_stamp()
        signature_data = method + timestamp + endpoint + payload
        message = bytes(signature_data, 'utf-8')
        secret = bytes(self.api_secret, 'utf-8')
        hash = hmac.new(secret, message, hashlib.sha256)
        return hash.hexdigest(), timestamp
    
    def get_product_id(self,symbol):
    
        url = f"https://api.delta.exchange/v2/products/{symbol}"
        response = requests.get(url, headers= self.headers)

        if response.status_code == 200:
            data = response.json()
            if data["success"]:
                print(data["result"]["id"])
                return data["result"]["id"]   
            else:
                return "Could not retrieve data. Please check the product symbol."
        else:
            return f"Failed to retrieve data. HTTP Status code: {response.status_code}"
        
            
    def place_order(self,side, qty, product_id, order_type="market_order", price=None):
        method = 'POST'
        url = "https://api.delta.exchange/v2/orders"
        endpoint = "/v2/orders"
        query_string = ''

        params = {
            "order_type": order_type,
            "size": qty,
            "side": side,
            "product_id": product_id,
        }
        if price:
            params["limit_price"] = str(price)

        payload = json.dumps(params).replace('','')
        signature, timestamp = self.Place_generate_signature(method, endpoint, payload)

        headers = {
            'api-key': self.api_key,
            'timestamp': timestamp,
            'signature': signature,
            'User-Agent': 'rest-client',
            'Content-Type': 'application/json'
        }

        response = requests.post(url, data=payload, headers=headers)
        return response.json()
    
    
    def add_to_excel(self , timestamp , entry_price , exit_price, stop_loss, take_profit):
        try:
              wb = load_workbook('BUY.xlsx')
              sheet = wb.active
        except FileNotFoundError:
              wb = Workbook()
              sheet = wb.active
              sheet.append(['Entry Time', 'Entry Price', 'Exit Price','stop_loss','take_profit'])

        for row in sheet.iter_rows(min_row=2, max_col=7):
            existing_entry = [cell.value for cell in row]
            new_entry = [entry_price, exit_price, stop_loss, take_profit]

            if existing_entry[1:] == new_entry:
                print("Duplicate entry found. Exiting without adding.")
                return

        sheet.append([timestamp, entry_price, exit_price,stop_loss,take_profit])
        wb.save('Trades.xlsx')
        print("Entry added successfully.")
    
    
    def Strategy(self,close_data):
        current_price = int(close_data)
        time_stamp = self.livedf['candel_start'].iloc[-1]
        if len(self.livedf) >= 2 and len(self.historydf) >= 2:
            if not pd.isna(self.livedf['current_EMA'].iloc[-2]) and not pd.isna(self.livedf['current_SMA'].iloc[-2]):
                diff = abs(self.livedf['current_SMA'].iloc[-1] - self.livedf['current_EMA'].iloc[-1])
                if (self.livedf['current_EMA'].iloc[-2] < self.livedf['current_SMA'].iloc[-2]) and (diff <= 1):
                    print("##################")
                    print(" BUY Signal ")
                    print("##################")
                    self.entry = float(self.livedf['close'].iloc[-1])
                    self.stoploss= int(self.entry  - 100)
                    self.takeprofit = int(self.entry + 200)
                    self.exit = self.stoploss
                    print(f"The current timestamp{self.livedf['candel_start'].iloc[-1]}")
                    print(f"The entry price is {self.entry}")
                    print(f"The  price is exit  {self.exit }")
                    print(f"The stoploss price is {self.stoploss}")

               # crossdown Strategy    
            if not pd.isna(self.livedf['EMA'].iloc[-2]) and not pd.isna(self.livedf['middleband'].iloc[-2]):
                diff = abs(self.livedf['middleband'].iloc[-1] - self.livedf['EMA'].iloc[-1])
                if (self.livedf['EMA'].iloc[-2] > self.livedf['middleband'].iloc[-2]) and (diff <= 1):
                    print("##################")
                    print(" SELL Signal ")
                    print("##################")
                    self.entry = float(self.livedf['close'].iloc[0])
                    self.stoploss= int(self.entry  + 100)
                    self.takeprofit = int(self.entry - 200)
                    self.exit = self.stoploss
                    print(f"The current timestamp{self.livedf['close'].iloc[-1]}")
                    print(f"The entry price is {self.entry}")
                    print(f"The  price is exit  {self.exit }")
                    print(f"The stoploss price is {self.stoploss}")

        if current_price == self.stoploss:
            print("The trade ended  with a loss")
            # Setting the exit to stoploss
            self.exit = self.stoploss
            #add the data to excel 
            self.add_to_excel(time_stamp,self.entry,self.exit,self.stoploss,self.takeprofit)

        if current_price == self.takeprofit:
            print("The trade ended  with a profit")
            # Setting the exit to the take profit
            self.exit = self.takeprofit
            #add the data to excel 
            self.add_to_excel(time_stamp,self.entry,self.exit,self.stoploss,self.takeprofit)
        

    
    def calculate_ema(self, prices, period):
        multiplier = 2 / (period + 1)
        return prices.ewm(alpha=multiplier, adjust=False).mean()
    
    def LiveEMA_Middle_band(self, close_data,start_time):
        # Assuming self.finaldf is a DataFrame with 'EMA' and 'middleband' columns
        Previous_EMA = self.finaldf['EMA'].iloc[-1]
        Previous_middle = self.finaldf['middleband'].iloc[-1]
        span = 5
        multiplier = 2 / (span + 1)
        current_EMA = (close_data - Previous_EMA) * multiplier + Previous_EMA

        # SMA Calculation
        sma_span = 15  # Define the period for the SMA calculation
        # To calculate the SMA, fetch the last (sma_span-1) closing prices and include the new close_data
        if len(self.finaldf['close']) >= sma_span - 1:
            recent_closes = self.finaldf['close'].tail(sma_span - 1).tolist() + [close_data]
            current_SMA = sum(recent_closes) / sma_span
        else:
            # If there's not enough historical 'close' data, calculate SMA with available data plus the new close_data
            recent_closes = self.finaldf['close'].tolist() + [close_data]
            current_SMA = sum(recent_closes) / len(recent_closes)


        live_EMA = pd.DataFrame({ 
            'candel_start':[pd.to_datetime(start_time / 1000000, unit='s')],
            'close': [close_data],
            'Previous_EAM': [Previous_EMA],
            'Previous_middle': [Previous_middle],
            'current_EMA': [current_EMA],
            'current_SMA':[current_SMA]
        })
        self.livedf = pd.concat([self.livedf,live_EMA],ignore_index=True)
        self.Strategy(close_data)


    def Update_data(self,data):
        if data['close'] is None or data['candle_start_time'] is None:
            print("Warning: close or candle_start_time value is None. Skipping Update_data.")
            return
        start_time = data['candle_start_time']
        open_data = float(data['open'])
        close_data = float(data['close'])
        high_data = float(data['high'])
        low_data = float(data['low'])
        volume_data = float(data['volume'])

        if np.isnan(close_data):  # Check if close_data is NaN
            print("Warning: close_data is NaN. Skipping Update_data.")
            return
        # print(f"Previous:{self.historydf['time'].iloc[-2]}")
        # print(f"Current:{self.historydf['time'].iloc[-1]}")
        try:
            new_df = pd.DataFrame({
                    'time': [pd.to_datetime(start_time / 1000000, unit='s')],
                    'open': [open_data],
                    'close': [close_data],
                    'high': [high_data],
                    'low': [low_data],
                    'volume': [volume_data]
                })
            self.historydf = pd.concat([self.historydf, new_df], ignore_index=True)
            if self.historydf['time'].iloc[-1] != self.historydf['time'].iloc[-2]:
                pre_closing =self.historydf['close'].iloc[-2]
                pre_open =self.historydf['open'].iloc[-2]
                pre_time =self.historydf['time'].iloc[-2]
                pre_low =self.historydf['low'].iloc[-2]
                pre_high =self.historydf['high'].iloc[-2]
                pre_volume =self.historydf['volume'].iloc[-2]
                pre_new_df =pd.DataFrame({
                    'time': [pre_time],
                    'open': [pre_open],
                    'close':[pre_closing],
                    'high': [pre_high],
                    'low':  [pre_low],
                    'volume': [pre_volume]
                })
                self.finaldf = pd.concat([self.finaldf, pre_new_df], ignore_index= True)
                self.finaldf['EMA'] = round(self.calculate_ema(self.finaldf['close'],5),1)
                self.finaldf['middleband'] = self.finaldf['close'].rolling(window=15).mean().round(1)
                self.finaldf['upper'] = 0 
                self.finaldf['lower'] = 0
                # print(self.finaldf.tail(1))
                self.finaldf.to_excel("EMA.xlsx")
            self.LiveEMA_Middle_band(close_data,start_time)
            
        except ValueError as e:
            print(f"Error converting data types: {e}")
            return

    def on_message(self,ws, message):
        data = json.loads(message)
        # print(data)
        self.Update_data(data)

    def on_close(self,ws, close_status_code, close_msg):
        print("Websocket closed")

    def on_error(self,ws, error):
        print(f"Error: {error}")

    def on_open(self,ws):
        print("Websocket opened")
        print("Waiting for the Signal....")
    # Send authentication message
        timestamp = self.get_time_stamp()
        signature_data = f"GET{timestamp}/live"
        signature = self.generate_signature( self.api_secret, signature_data)

        auth_msg = {
            "type": "auth",
            "payload": {
                "api-key": self.api_key,
                "signature": signature,
                "timestamp": timestamp
            }
        }

        ws.send(json.dumps(auth_msg))  
        # Subscribe to channels
        ''' 
        if u want to chang the candle_stick time u can  change candlestick_1m for example:
        you want to 15min time framse  u chang it to candlestick_15m

        '''
        subscribe_msg = {
            "type": "subscribe",
            "payload": {
                "channels": [
                    {"name": "candlestick_1m", "symbols": ['BTCUSDT']},
                ]
            }
        }
        ws.send(json.dumps(subscribe_msg))


    def live_data(self):
      ws = websocket.WebSocketApp('wss://socket.delta.exchange', on_open=self.on_open, on_message=self.on_message, on_error=self.on_error, on_close=self.on_close)
      ws.run_forever()

api_key = 'VraFQTfHWk1w7Id1uC1pJvyKQ5AWy5'

api_secret = 'XJofp39iR4p7092qkzJdH9VPmoZmHtWv2x0huuJWscPerHtGqcf2Uo996JMj'


delta = Delat(api_key,api_secret,25)

delta.History()
delta.live_data()