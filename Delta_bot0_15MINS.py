import requests
import pandas as pd
import websocket
import datetime
import json
from talib import abstract
import hashlib
import hmac
import numpy as np
from openpyxl import Workbook, load_workbook
import time
import pyotp
import socket



class Delat():
    def __init__(self,api_key, api_secret,lev):
        self.api_key = api_key
        self.api_secret = api_secret
        self.leverage = lev
        self.EMA = 5
        self.BBL = 20
        self.BBSD = 3.0
        self.headers = {'Accept': 'application/json'}
        self.historydf = None
        self.update_data_df= None
        self.entry = 0
        self.exit = 0
        self.stoploss= 0
        self.takeprofit = 0
        self.crossover = False
        self.finaldf= None
        self.livedf= None
        self.max_reconnect_attempts = 1000
        self.reconnect_attempts = 0
        self.reconnect_delay = 1
        self.signal_type = 'none'
        self.pre5diff = False
        self.qty = 2
        self.half_qty = int(self.qty/2)
        self.pid =self.get_product_id('BTCUSDT')
        self.delay = 1
        self.userid=99288206
        self.orderdf = None
        self.df_Data = ()
        self.processed_ids = set()
        self.take_vale = 500
        self.stop_vale = 500



    def History(self):
        current_timestamp = int(datetime.datetime.now().timestamp())
        last_100_candel = current_timestamp - 90000

        params = {
            'resolution': '15m',
            'symbol': 'BTCUSDT',
            'start':last_100_candel, 
            'end': current_timestamp   
        }

        r = requests.get('https://api.delta.exchange/v2/history/candles', params=params, headers=self.headers)
        if r.status_code == 200:
            json_data = r.json()
            df = pd.DataFrame(json_data['result'])
            df['time'] = pd.to_datetime(df['time'], unit='s')  
            # df['time'] = df['time'].dt.tz_localize('UTC').dt.tz_convert('Asia/Kolkata').dt.strftime('%Y-%m-%d %H:%M:%S')
            self.historydf = df
            self.historydf = self.historydf.iloc[::-1].reset_index(drop=True)
            # self.historydf = self.historydf.drop(self.historydf.index[-1])
            
            self.historydf['SMA'] = np.round(abstract.SMA(self.historydf['close'], timeperiod=self.BBL),1)
           
            self.historydf['EMA'] =np.round(abstract.EMA(self.historydf['close'],self.EMA),1)
           
            self.finaldf = self.historydf
            self.livedf = self.historydf
            self.save_history_to_excel('history_data.xlsx')

        else:
            print(f"Failed to fetch data. Status code: {r.status_code}")
    


    def save_history_to_excel(self, file_path='history_data.xlsx'):
        try:
            # Using the 'openpyxl' engine to enable writing to Excel files
            self.historydf.to_excel(file_path, engine='openpyxl', index=False)
            print(f"History DataFrame saved to {file_path}")
            
        except Exception as e:
            print(f"Failed to save DataFrame to Excel. Error: {e}")


    def add_to_df(self,responce,take_profit):

        order_data = responce["result"]
        filter_data = {
            "order_id": order_data.get("id",""),
            "product_symbol": order_data.get("product_symbol",""),
            "size":order_data.get("size",0.0),
            "side":order_data.get("side",""),
            "commission":order_data.get("commission",""),
            "Entry_price" :order_data.get("average_fill_price",""),
            "Take_profit":take_profit
        }
        self.df_Data= tuple(filter_data.values())
        print(self.df_Data)

        columns = ["order_id","product_symbol","size","side","commission", "Entry_price","Take_profit"]
        new_order = pd.DataFrame([self.df_Data],columns=columns)

        if self.orderdf is None:
            self.orderdf = new_order
        else:
            self.orderdf = pd.concat([self.orderdf, new_order], ignore_index=True)
            print("The data added done")
        self.orderdf.to_excel("orders_details.xlsx")
        print(self.orderdf)

    def read_excel_file(self):
        data = pd.read_excel("orders_details.xlsx")
        get_last_data = data.iloc[-1]
        return get_last_data
    
    def generate_signature(self,secret, message):
        message = bytes(message, 'utf-8')
        secret = bytes(secret, 'utf-8')
        hash = hmac.new(secret, message, hashlib.sha256)
        return hash.hexdigest()

    def get_time_stamp(self):
        d = datetime.datetime.utcnow()
        epoch = datetime.datetime(1970, 1, 1)
        return str(int((d - epoch).total_seconds()))
    
    def Place_generate_signature(self,method, endpoint, payload):
        timestamp = self.get_time_stamp()
        signature_data = method + timestamp + endpoint + payload
        message = bytes(signature_data, 'utf-8')
        secret = bytes(self.api_secret, 'utf-8')
        hash = hmac.new(secret, message, hashlib.sha256)
        return hash.hexdigest(), timestamp
    
    def get_product_id(self,symbol):
        url = f"https://api.delta.exchange/v2/products/{symbol}"
        response = requests.get(url, headers= self.headers)
        if response.status_code == 200:
            data = response.json()
            if data["success"]:
                print(data["result"]["id"])
                return data["result"]["id"]   
            else:
                return "Could not retrieve data. Please check the product symbol."
        else:
            return f"Failed to retrieve data. HTTP Status code: {response.status_code}"
        

    def Get_wallet_info(self):
        method = 'GET'
        endpoint = '/v2/wallet/balances'
        url  = 'https://api.delta.exchange/v2/wallet/balances'
        signature, timestamp = self.Place_generate_signature(method,endpoint,'')
        headers = {'Accept': 'application/json',
                    'api-key': self.api_key,
                    'signature': signature,
                    'timestamp': timestamp}
        response = requests.get(url, headers=headers)
        # print(response.status_code)
        # print(response.text)
        return response.json()
    

    def set_Leverage(self,product_id,leverage):
        method= 'POST'
        endpoint =f'/v2/products/{product_id}/orders/leverage'
        url = f'https://api.delta.exchange/v2/products/{product_id}/orders/leverage'
        params = {"leverage":leverage}
        payload = json.dumps(params).replace(' ', '')
        signature, timestamp = self.Place_generate_signature(method, endpoint, payload )
        headers = {
        'api-key': self.api_key,
        'timestamp': timestamp,
        'signature': signature,
        'Content-Type': 'application/json',
        'Accept': 'application/json'
        }
        response = requests.post(url, data=payload, headers=headers)
        print(response.status_code)
        print(response.text)
        return response.json()
    
    def Get_Positions(self,product_id=None, underlying_symbol=None):
        if not product_id and not underlying_symbol:
            print('Either product_id or underlying_symbol must be provided')
            return
        method= 'GET'
        endpoint = '/v2/positions'
        url = 'https://api.delta.exchange/v2/positions'
        params = {}

        if product_id:
            params['product_id'] = product_id
        if underlying_symbol:
            params['underlying_asset_symbol'] = underlying_symbol

        # payload = json.dumps(params).replace(' ', '')
 
        query_string = '?' + '&'.join(f'{key}={value}' for key, value in params.items()) if params else ''
        endpoint += query_string
        url += query_string
        signature, timestamp = self.Place_generate_signature(method, endpoint , '')
        headers = {
        'api-key': self.api_key,
        'timestamp': timestamp,
        'signature': signature,
        'User-Agent': 'rest-client',
        'Content-Type': 'application/json'
        }

        response = requests.get(url , headers= headers)
        return response.json()
    
    def close_all_positions(self, close_all_portfolio=False, close_all_isolated=True, user_id = 0):
        method = 'POST'
        endpoint = '/v2/positions/close_all'
        url = f"https://api.delta.exchange/v2/positions/close_all"
        params = {
            "close_all_portfolio": close_all_portfolio,
            "close_all_isolated": close_all_isolated,
            "user_id": user_id
        }
        payload = json.dumps(params).replace(' ', '')
        signature, timestamp = self.Place_generate_signature(method, endpoint, payload)
        headers = {
            'api-key': self.api_key,
            'timestamp': timestamp,
            'signature': signature,
            'User-Agent': 'rest-client',
            'Content-Type': 'application/json'
        }
        response = requests.post(url, data=payload, headers=headers)
        if response.status_code == 200:
            try:
                print(response.text)
                return response.json()
            except json.decoder.JSONDecodeError:
                print(f'Failed to parse JSON: {response.text}')
        else:
            print(f'Failed to close all positions: {response.status_code} - {response.text}')


    def place_bracket_order(self,side, qty, product_id, order_type="limit_order", price=None, stop_loss=None, take_profit=None):
        method = 'POST'
        url = "https://api.delta.exchange/v2/orders"
        endpoint = "/v2/orders"

        params = {
            "order_type": order_type,
            "size": qty,
            "side": side,
            "product_id": product_id,
        }

        if price:
            params["limit_price"] = str(price)

        if stop_loss:
            params["bracket_stop_loss_price"] = str(stop_loss)
            params["bracket_stop_loss_limit_price"] = str(stop_loss)  # Assuming limit price is the same as stop price

        if take_profit:
            params["bracket_take_profit_price"] = str(take_profit)
            params["bracket_take_profit_limit_price"] = str(take_profit)  # Assuming limit price is the same as take profit price

        payload = json.dumps(params).replace(' ', '')
        signature, timestamp = self.Place_generate_signature(method, endpoint, payload)

        headers = {
            'api-key': self.api_key,
            'timestamp': timestamp,
            'signature': signature,
            'User-Agent': 'rest-client',
            'Content-Type': 'application/json'
        }

        response = requests.post(url, data=payload, headers=headers)
        print(response.status_code)
        print(response.text)
        if response.status_code == 200:
            try:
                print(response.text)
                return response.json()
            except json.decoder.JSONDecodeError:
                print(f'Failed to parse JSON: {response.text}')
        else:
            print(f'Failed to close all positions: {response.status_code} - {response.text}')
    
    
  

    def place_order(self,side, qty, product_id, order_type="market_order", price=None):
        method = 'POST'
        url = "https://api.delta.exchange/v2/orders"
        endpoint = "/v2/orders"
        query_string = ''

        params = {
            "order_type": order_type,
            "size": qty,
            "side": side,
            "product_id": product_id,

        }
        if price:
            params["limit_price"] = str(price)

        payload = json.dumps(params).replace(' ','')
        signature, timestamp = self.Place_generate_signature(method, endpoint, payload)

        headers = {
            'api-key': self.api_key,
            'timestamp': timestamp,
            'signature': signature,
            'User-Agent': 'rest-client',
            'Content-Type': 'application/json'
        }

        response = requests.post(url,data=payload, headers=headers)
        if response.status_code == 200:
            try:
                print(response.text)
                return response.json()
            except json.decoder.JSONDecodeError:
                print(f'Failed to parse JSON: {response.text}')
        else:
            print(f'Failed to place Buy: {response.status_code} - {response.text}')

    
    
    def add_to_excel(self, timestamp, entry_price, signal):
        try:
            # Load existing workbook or create a new one if not exists
            wb = load_workbook('Trades.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(['Entry Time', 'Entry Price', 'Signal'])

        new_entry = [timestamp, entry_price, signal]

        # Check for duplicate entry
        if sheet.max_row > 1:  
            # Check if there is at least one entry
            # Access the last row in the Excel sheet
            last_row = sheet[sheet.max_row]
            # Extract values from the last row cells
            existing_entry = [cell.value for cell in last_row[:3]]
            
            
            if existing_entry[2] == new_entry[2]:
                print("Duplicate entry found. Exiting without adding.")
                return True

        # Append trade details to the Excel sheet
        sheet.append([timestamp, entry_price, signal])
        # Save workbook
        wb.save('Trades.xlsx')
        print("Entry added successfully.")
            
    def check_for_new_entries(self,df, processed_set):
        if self.orderdf is not None:
    # Check for entries not in the processed set
            new_entries = df[~df['order_id'].isin(processed_set)]
            if not new_entries.empty:
                print("New entries found:", new_entries)
                # Add new IDs to the set
                processed_set.update(new_entries['order_id'])
                return True
        else:
            pass


    def take_profit(self):
        pos = self.Get_Positions(139)
        new_entry = self.check_for_new_entries(df=self.orderdf,processed_set=self.processed_ids)
        entry_price = pos["result"]["entry_price"]
        if entry_price is not None:
            if new_entry is True:
                print("HEllO")
                # Assuming self.orderdf is properly initialized and contains 'side' column
                last_side = self.orderdf["side"].iloc[-1].upper() 
                     # Convert to upper case to handle case variations
                print("The Last side ",last_side)
                if last_side == 'BUY':
                        self.place_order(side='sell', qty=self.half_qty, product_id=139, order_type='limit_order', price=self.orderdf["Take_profit"].iloc[-1])
                        print("Placing the BUY side Take profit")
                elif last_side == 'SELL':
                        print(self.half_qty)
                        self.place_order(side='buy', qty=self.half_qty, product_id=139, order_type='limit_order', price=self.orderdf["Take_profit"].iloc[-1])
                        print("Placing the SELL side Take profit")
            else:
                pass

    
    def Strategy(self,close_data): 

        self.finaldf['Diff'] = abs(self.finaldf['EMA'] - self.finaldf['SMA'])       
        last_index = self.finaldf.index[-2]
        start_index = max(0, last_index - 7)
        self.finaldf.to_excel('EMA&SMA.xlsx')
        if (self.finaldf['Diff'].iloc[start_index:last_index + 1] < 50).any():
            self.pre5diff = True
            #print (self.pre5diff)                                                             #edited
        current_price = int(close_data)
        time_stamp = self.livedf['candel_start'].iloc[-1]
        self.livedf['diff'] = abs(self.livedf['current_SMA'].iloc[-1] - self.livedf['current_EMA'].iloc[-1]) 

        # print(self.livedf.iloc[-1])             

        # print(len(self.livedf))
        # print(len(self.historydf))
        # print(len(self.finaldf) )
        if len(self.livedf) >= 5 and len(self.historydf) >= 5 and len(self.finaldf) >= 5:

            if(self.pre5diff == True):

                if(self.livedf['diff'].iloc[-1] >= 100 and self.livedf['diff'].iloc[-1] <=120):

                    if (self.livedf['current_EMA'].iloc[-1] > self.livedf['current_SMA'].iloc[-1]):

                        print("##################")
                        print(" BUY Signal ")
                        print("##################")
                        self.signal_type = 'buy'                                                        #need edit in the stratagy
                        self.entry = float(self.livedf['close'].iloc[-1])
                        self.stoploss = self.entry - self.stop_vale
                        print(f"The current timestamp{self.livedf['candel_start'].iloc[-1]}")
                        print(f"The entry price is {self.entry}")

                        Duplicate = self.add_to_excel(time_stamp,self.entry,self.signal_type)
                        pos= self.Get_Positions(139)
                       
                        # Buy order placement
                        if pos["result"]["entry_price"] is None:
                            responce = self.place_bracket_order(side='buy',qty=self.qty, product_id=139 ,order_type="market_order",stop_loss=self.stoploss)
                            takeprofit =float(responce["result"]["average_fill_price"]) + self.take_vale
                            self.add_to_df(responce,takeprofit)
                            res= self.place_order(side='sell',qty=self.half_qty,product_id=139,order_type='limit_order',price=takeprofit)
                    
                            print(res)
                            print("placing order two of take profit BUY ")

                        elif self.orderdf["side"].iloc[-1] != self.signal_type:
                            self.close_all_positions(close_all_portfolio=True,user_id=self.userid)
                            responce = self.place_bracket_order(side='buy',qty=self.qty, product_id=139 ,order_type="market_order",stop_loss=self.stoploss)
                        
                            takeprofit =float(responce["result"]["average_fill_price"]) + self.take_vale
                            self.add_to_df(responce,takeprofit)

                            self.place_order(side='sell', qty=self.half_qty , product_id=139, order_type='limit_order',price=takeprofit)

                        else:
                            print("Order is Repeating")

                        # if pos is  None 
                            #place order (BUY)
                            #place bracket order 
                        #else  
                            #close all pos
                            #palce order 
                            #place bracket order   



                    if (self.livedf['current_EMA'].iloc[-1] < self.livedf['current_SMA'].iloc[-1]):
                        print("##################")
                        print(" SELL Signal ")
                        print("##################")
                        self.signal_type = 'sell'
                        self.entry = float(self.livedf['close'].iloc[-1])
                        self.takeprofit =  self.entry - self.take_vale
                        self.stoploss = self.entry + self.stop_vale
                        print(f"The current timestamp{self.livedf['candel_start'].iloc[-1]}")
                        print(f"The entry price is {self.entry}")
                        Duplicate= self.add_to_excel(time_stamp,self.entry,self.signal_type)
                        pos= self.Get_Positions(139)
                        # sell order place
                        if pos["result"]["entry_price"] is None:
                            responce= self.place_bracket_order(side='sell',qty=self.qty,product_id=139 ,order_type="market_order",stop_loss=self.stoploss)

                            takeprofit =float(responce["result"]["average_fill_price"]) - self.take_vale
                            self.add_to_df(responce,takeprofit)

                            self.place_order(side='buy',qty=self.half_qty,product_id=139,order_type='limit_order',price=takeprofit)
            
                            print("placing order two of take profit SELL")

                        elif  self.orderdf["side"].iloc[-1] != self.signal_type:
                            self.close_all_positions(close_all_portfolio=True,user_id=self.userid)
                            responce= self.place_bracket_order(side='sell',qty=self.qty,product_id=139 ,order_type="market_order",stop_loss=self.stoploss)
                            takeprofit =float(responce["result"]["average_fill_price"]) - self.take_vale
                            self.add_to_df(responce,takeprofit)
                            self.place_order(side='buy', qty=self.half_qty , product_id=139,order_type ='limit_order', price=takeprofit)
                    print("trade taken after 100 diff ")

          
            elif (self.livedf['diff'].iloc[-1]  <= 5):                                                                        #edited
                        # crossdown Strategy  
                if not pd.isna(self.livedf['current_EMA'].iloc[-2]) and not pd.isna(self.livedf['current_SMA'].iloc[-2]):
                    if (self.finaldf['EMA'].iloc[-3] > self.finaldf['SMA'].iloc[-3]):
                        print("##################")
                        print(" SELL Signal ")
                        print("##################")
                        self.signal_type = 'sell'
                        self.entry = float(self.livedf['close'].iloc[-1])
                        self.takeprofit =  self.entry - self.take_vale
                        self.stoploss = self.entry + self.stop_vale
                        print(f"The current timestamp{self.livedf['candel_start'].iloc[-1]}")
                        print(f"The entry price is {self.entry}")
                        Duplicate= self.add_to_excel(time_stamp,self.entry,self.signal_type)    
                        pos= self.Get_Positions(139)
                        if pos["result"]["entry_price"] is None:
                            responce1 = self.place_bracket_order(side='sell',qty=self.qty,product_id=139 ,order_type="market_order",stop_loss=self.stoploss)
                            takeprofit =float(responce1["result"]["average_fill_price"]) - self.take_vale
                            self.add_to_df(responce1,takeprofit)
                            self.place_order(side='buy',qty=self.half_qty,product_id=139,order_type='limit_order',price=takeprofit)
        
                            print("placing order two of take profit SELL ")

                        elif  self.orderdf["side"].iloc[-1]!= self.signal_type:
                            self.close_all_positions(close_all_portfolio=True,user_id=self.userid)
                            responce1 = self.place_bracket_order(side='sell',qty=self.qty,product_id=139 ,order_type="market_order",stop_loss=self.stoploss)

                            takeprofit = float(responce1["result"]["average_fill_price"]) - self.take_vale
                            self.add_to_df(responce1,takeprofit)
                            self.place_order(side='buy', qty=self.half_qty , product_id=139,order_type='limit_order', price=takeprofit)


                    # crossup Strategy    
                if not pd.isna(self.livedf['current_EMA'].iloc[-2]) and not pd.isna(self.livedf['current_SMA'].iloc[-2]):
                    if (self.finaldf['EMA'].iloc[-3] < self.finaldf['SMA'].iloc[-3]):
                            print("##################")
                            print(" BUY Signal ")
                            print("##################")
                            self.signal_type = 'buy'
                            self.entry = float(self.livedf['close'].iloc[-1])
                            self.stoploss = self.entry - self.stop_vale
                            print(f"The current timestamp{self.livedf['candel_start'].iloc[-1]}")
                            print(f"The entry price is {self.entry}")
                            Duplicate= self.add_to_excel(time_stamp,self.entry,self.signal_type)

                            pos= self.Get_Positions(139)
                            # Buy order placement
                            if pos["result"]["entry_price"] is None:
                                responce = self.place_bracket_order(side='buy',qty=self.qty,product_id=139 ,order_type="market_order",stop_loss=self.stoploss)
                                takeprofit =float(responce["result"]["average_fill_price"]) + self.take_vale

                                self.add_to_df(responce,takeprofit)
                                res= self.place_order(side='sell',qty=self.half_qty,product_id=139,order_type='limit_order',price = takeprofit)

                                print("this is responce",res)
                                print("placing order two of take profit  BUY ")
                            elif self.orderdf["side"].iloc[-1]!= self.signal_type:
                                self.close_all_positions(close_all_portfolio=True,user_id=self.userid)
                                responce1=self.place_bracket_order(side='buy',qty=self.qty,product_id=139 ,order_type="market_order",stop_loss=self.stoploss)
                                takeprofit =float(responce1["result"]["average_fill_price"]) + self.take_vale
                                self.add_to_df(responce1,takeprofit)
                                self.place_order(side='sell',qty=self.half_qty,product_id=139,order_type='limit_order',price=takeprofit)
                                
                print("trade taken while crossing ") 
        # self.take_profit()
             

    
    def calculate_ema(self, prices, period):
        multiplier = 2 / (period + 1)
        return prices.ewm(alpha=multiplier, adjust=True).mean()
    
    def Live_EMA_SMA(self, close_data, start_time):
        # Assuming self.finaldf is a DataFrame with 'EMA' and 'SMA' columns
        Previous_EMA = self.finaldf['EMA'].iloc[-1]
        Previous_SMA = self.finaldf['SMA'].iloc[-1]
        span = self.EMA
        multiplier = 2 / (span + 1)
        current_EMA = (close_data - Previous_EMA) * multiplier + Previous_EMA

        # SMA Calculation
        sma_span = self.BBL  # Define the period for the SMA calculation
        # To calculate the SMA, fetch the last (sma_span-1) closing prices and include the new close_data
        if len(self.finaldf['close']) >= sma_span - 1:
            recent_closes = self.finaldf['close'].tail(sma_span - 1).tolist() + [close_data]
            current_SMA = sum(recent_closes) / sma_span
        else:
            # If there's not enough historical 'close' data, calculate SMA with available data plus the new close_data
            recent_closes = self.finaldf['close'].tolist() + [close_data]
            current_SMA = sum(recent_closes) / len(recent_closes)


        live_EMA = pd.DataFrame({ 
            'candel_start':[pd.to_datetime(start_time / 1000000, unit='s')],
            'close': [close_data],
            'Previous_EMA': [Previous_EMA],
            'Previous_SMA': [Previous_SMA],
            'current_EMA': [current_EMA],
            'current_SMA':[current_SMA]
  
        })
        
        self.livedf = pd.concat([self.livedf,live_EMA],ignore_index=True)                           # only DF with SMA an EMA need to be added to self.livedf
        #print(self.livedf)
        #print (self.livedf)
        self.Strategy(close_data)



    def Update_data(self,data):

        if data['close'] is None or data['candle_start_time'] is None:
            print("Warning: close or candle_start_time value is None. Skipping Update_data.")
            return
        start_time = data['candle_start_time']
        open_data = float(data['open'])
        close_data = float(data['close'])
        high_data = float(data['high'])
        low_data = float(data['low'])
        volume_data = float(data['volume'])

        if np.isnan(close_data):  # Check if close_data is NaN
            print("Warning: close_data is NaN. Skipping Update_data.")
            return
        #print(f"Previous:{self.historydf['time'].iloc[-2]}")
        #print(f"Current:{self.historydf['time'].iloc[-1]}")
        try:
            new_df = pd.DataFrame({
                    'time': [pd.to_datetime(start_time / 1000000, unit='s')],
                    'open': [open_data],
                    'close': [close_data],
                    'high': [high_data],
                    'low': [low_data],
                    'volume': [volume_data]
                })
            self.historydf = pd.concat([self.historydf, new_df], ignore_index=True)
            if self.historydf['time'].iloc[-1] != self.historydf['time'].iloc[-2]:
                pre_closing =self.historydf['close'].iloc[-2]
                pre_open =self.historydf['open'].iloc[-2]
                pre_time =self.historydf['time'].iloc[-2]
                pre_low =self.historydf['low'].iloc[-2]
                pre_high =self.historydf['high'].iloc[-2]
                pre_volume =self.historydf['volume'].iloc[-2]
                pre_new_df =pd.DataFrame({
                    'time': [pre_time],
                    'open': [pre_open],
                    'close':[pre_closing],
                    'high': [pre_high],
                    'low':  [pre_low],
                    'volume': [pre_volume]
                })
                self.finaldf = pd.concat([self.finaldf, pre_new_df], ignore_index= True)
                self.finaldf['EMA'] = round(self.calculate_ema(self.finaldf['close'],self.EMA),1)
                self.finaldf['SMA'] = self.finaldf['close'].rolling(window=self.BBL).mean().round(1)    #edited
            self.Live_EMA_SMA(close_data,start_time)
            
        except ValueError as e:
            print(f"Error converting data types: {e}")
            return
        

    def send_heartbeat(self,ws):
        heartbeat={
        "type": "enable_heartbeat"
        }
        ws.send(json.dumps(heartbeat))
        self.Add_IP(email="ajay17joe@gmail.com",password="Rhino1499#",mfa='ALUM267K6BMUUPYU')
         
    def on_message(self,ws, message):
        data = json.loads(message)
        self.Update_data(data)

    def on_close(self,ws, close_status_code, close_msg):
        print("Websocket closed")

        if self.reconnect_attempts < self.max_reconnect_attempts:
            print(f"Attempting to reconnect, attempt {self.reconnect_attempts + 1}")
            time.sleep(self.reconnect_delay)  # Wait before attempting to reconnect
            self.reconnect_attempts += 1  # Increment the reconnection attempt counter
            self.Add_IP(email="ajay17joe@gmail.com",password="Rhino1499#",mfa='ALUM267K6BMUUPYU')
            pid= self.get_product_id('BTCUSDT')
            self.Get_wallet_info()
            self.set_Leverage(pid,50)
            self.History()
            self.live_data()# Attempt to reconnect
        else:
            print("Maximum reconnection attempts reached. Not attempting to reconnect.")

    def on_error(self,ws, error):
        er = str(error)
        if 'close' in er :
            pass
        else:
            print(f"Error: {error}")
    
    def on_open(self,ws):
        print("Websocket opened")
        print("Waiting for the Signal....")
    # Send authentication message
        timestamp = self.get_time_stamp()
        signature_data = f"GET{timestamp}/live"
        signature = self.generate_signature( self.api_secret, signature_data)

        auth_msg = {
            "type": "auth",
            "payload": {
                "api-key": self.api_key,
                "signature": signature,
                "timestamp": timestamp
            }
        }

        ws.send(json.dumps(auth_msg))  
        # Subscribe to channels
        ''' 
        if u want to chang the candle_stick time u can  change candlestick_1m for example:
        you want to 15min time framse  u chang it to candlestick_15m

        '''
        subscribe_msg = {
            "type": "subscribe",
            "payload": {    
                "channels": [
                    {"name": "candlestick_15m", "symbols": ['BTCUSDT']},
                ]
            }
        }
        ws.send(json.dumps(subscribe_msg))
        self.send_heartbeat(ws)

    def live_data(self):
      ws = websocket.WebSocketApp('wss://socket.delta.exchange', on_open=self.on_open, on_message=self.on_message, on_error=self.on_error, on_close=self.on_close)
      ws.run_forever()

    def Add_IP(self,email, password,mfa):
        hostname = socket.gethostname()
        ipv4 = socket.gethostbyname(hostname)
        # ipv6 = socket.getaddrinfo(hostname,0,socket.AF_INET6)[1][4][0]
        topt = pyotp.TOTP(mfa)
        otp = topt.now()
        get_info = self.Get_wallet_info()
        ip_to_add = None
        print(get_info)
        if get_info['success'] is True :
            pass
        else:
            ip_to_add = get_info["error"]["client_ip"]
            print(topt.verify(otp))
            print(otp)
            
            url = 'http://cdn.deltaex.org/v2/api_keys/whitelist_ip'
            parms  = {
                    "email":email,
                    "password": password,
                    "mfa_code": otp,
                    "api_key": self.api_key,
                    "ip_addresses":[ip_to_add]
                }
            print(parms)
            paylord = json.dumps(parms)
            headers = {
                    'Accept': 'application/json',
                    'Content-Type': 'application/json' 
                }
            if topt.verify(otp):
                response = requests.post(url=url,data=paylord,headers=headers)
                print(response.status_code)
                print(response.text)
            else:
                self.Add_IP(email="ajay17joe@gmail.com",password="Rhino1499#",mfa='ALUM267K6BMUUPYU')

# api_key = 'VraFQTfHWk1w7Id1uC1pJvyKQ5AWy5'
# api_secret = 'XJofp39iR4p7092qkzJdH9VPmoZmHtWv2x0huuJWscPerHtGqcf2Uo996JMj'

# api_key = 'oNhfyfYF7JWO8Dg49e5K9SzyYur3a3'
# api_secret = 'QtqmRXJlA7QzlfwnTxvPImavoO6excynkGxCbNNkpFwJugooy5s2kFiRhfPz'

api_key = 'PifFnRLv7RINsedJ0UyhWtfWj6qNLL'
api_secret = 'o27aiNsPFuboGmiSqYxIq6CbgkDliDFY2krX2H7SIRK3Smlv86uFNGF1CblC'


def get_symbol_price(symbol):
    headers = {
        'Accept': 'application/json',
    }

    r = requests.get('https://api.delta.exchange/v2/tickers', params={}, headers=headers)
    r = r.json()

    try:
        r = r['result']
    except:
        print('Price retrieval error')
        return get_symbol_price(symbol)  # Recursive call to try again

    symbol_price = next((i['close'] for i in r if i['symbol'] == symbol.upper()), None)

    if symbol_price is None:
        print(f"Could not find price for symbol: {symbol}")
        return None

    return symbol_price

symbol = 'BTCUSDT'
delta = Delat(api_key,api_secret,25)
delta.Add_IP(email="ajay17joe@gmail.com",password="Rhino1499#",mfa='ALUM267K6BMUUPYU')
# pid= delta.get_product_id(symbol)
# delta.Get_wallet_info()
# delta.set_Leverage(pid,50)
# delta.History()
# delta.live_data()
delta.read_excel_file()


# responce = {"meta":{},"result":{"cancellation_reason":None,"order_type":"limit_order","bracket_take_profit_price":None,"bracket_order":None,"reduce_only":"false","bracket_trail_amount":None,"side":"sell","stop_price":None,"bracket_stop_loss_limit_price":"63388","state":"closed","time_in_force":"gtc","unfilled_size":0,"quote_size":None,"client_order_id":None,"id":3493626021,"user_id":99288206,"bracket_take_profit_limit_price":None,"average_fill_price":"63288","mmp":"disabled","stop_order_type":None,"commission":"0","bracket_stop_loss_price":"63388","created_at":"2024-04-30T06:04:03.328135Z","paid_commission":"0.0759456","close_on_trigger":"false","product_id":139,"size":2,"limit_price":"63288","meta_data":{"cashflow":"0","ip":"2001:4490:4c35:ebb5:1d7e:4f35:38ba:4b17","otc":"False","pnl":"0","source":"api"},"product_symbol":"BTCUSDT","updated_at":"2024-04-30T06:04:03.559555Z","trail_amount":None,"stop_trigger_method":None},"success":"true"} 
# delta.add_to_df(responce,ltp-100)
# delta.take_profit()

# ltp = get_symbol_price(symbol)
# responce= delta.place_order(side='sell',qty=2,product_id=pid)
# print(responce["result"]["average_fill_price"])
# res= delta.place_bracket_order(side ='buy',qty=2,product_id=139,order_type="market_order",stop_loss=ltp-200)
# delta.place_order(side ='sell',qty=1,product_id=139,order_type='limit_order',price=ltp+300)  

# responce=delta.place_bracket_order(side ='sell',qty=1,product_id=139,stop_loss=ltp+100)
# print(responce)